import openpyxl
import json

wb_obj = openpyxl.load_workbook("item.xlsx")
wb_obj.active = wb_obj.sheetnames.index('enUS')
sheet_obj = wb_obj.active
item_map = {"None": "None"}
for row in sheet_obj.rows:
    item_map[row[2].value] = row[1].value

with open("../resources/item_en-US.json", "w") as f:
    json.dump(item_map, f)

wb_obj.active = wb_obj.sheetnames.index('zhCN')
sheet_obj = wb_obj.active
item_map = {"None": "无物品"}
for row in sheet_obj.rows:
    item_map[row[1].value] = row[0].value

with open("../resources/item_zh-CN.json", "w") as f:
    json.dump(item_map, f)